"""Build the monthly salary spreadsheet for the accountant.

One row per worker for a chosen month: days worked, hours (regular / 125% /
150% / total), the pay basis and rate, and the computed base / overtime / total
pay — everything the accountant needs to prepare payslips.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Role, Shift, User
from core.payroll import compute_payroll

# Column headers, in Hebrew (the accountant works in Hebrew).
HEADERS = [
    'שם העובד', 'ת.ז.', 'בסיס שכר', 'תעריף ₪', 'ימי עבודה',
    'שעות רגילות', 'שעות 125%', 'שעות 150%', 'סה"כ שעות',
    'שכר בסיס ₪', 'שעות נוספות ₪', 'סה"כ לתשלום ₪',
]
BASIS_HE = {'hourly': 'לפי שעה', 'daily': 'לפי יום', 'monthly': 'חודשי'}

_NAVY = 'FF14284B'
_HEAD_FILL = PatternFill('solid', fgColor=_NAVY)
_HEAD_FONT = Font(color='FFFFFFFF', bold=True, size=11)
_BORDER = Border(*(Side(style='thin', color='FFDDE3E9'),) * 4)


def _month_range(year, month):
    from django.utils import timezone
    now = timezone.localtime()
    start = now.replace(year=year, month=month, day=1, hour=0, minute=0,
                        second=0, microsecond=0)
    end = start.replace(year=year + (month == 12),
                        month=1 if month == 12 else month + 1)
    return start, end


def build_salary_excel(year, month):
    """Return the salary spreadsheet for `year`-`month` as PDF-like bytes (xlsx)."""
    start, end = _month_range(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}-{month:02d}'
    ws.sheet_view.rightToLeft = True

    ws.append(HEADERS)
    for c, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _BORDER

    workers = (User.objects.filter(role__in=[Role.WAREHOUSE, Role.DRIVER,
                                             Role.OFFICE, Role.MANAGER])
               .order_by('first_name', 'last_name'))
    for w in workers:
        shifts = Shift.objects.filter(worker=w, clock_in__gte=start,
                                      clock_in__lt=end, clock_out__isnull=False)
        p = compute_payroll(w, shifts)
        basis = w.pay_basis or 'hourly'
        if basis == 'daily':
            rate = float(w.daily_rate or 0)
        elif basis == 'monthly':
            rate = float(w.monthly_salary or 0)
        else:
            rate = float(w.hourly_rate or 0)
        ws.append([
            w.full_name, w.id_number or '', BASIS_HE.get(basis, basis), rate,
            p['days_worked'], p['regular_hours'], p['overtime_125_hours'],
            p['overtime_150_hours'], p['total_hours'],
            p['base_pay'], p['overtime_pay'], p['total_pay'],
        ])

    # Style the data rows and size the columns.
    last = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Money columns → 2 decimals.
    for col in (4, 10, 11, 12):
        for r in range(2, last + 1):
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    for c in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = \
            18 if c in (1,) else 13

    # A totals row for the money columns.
    if last >= 2:
        ws.append([])
        total_row = ws.max_row
        ws.cell(row=total_row, column=1, value='סה"כ').font = Font(bold=True)
        for col in (10, 11, 12):
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f'=SUM({letter}2:{letter}{last})')
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def salary_filename(year, month):
    return f'salary_{year}_{month:02d}.xlsx'
